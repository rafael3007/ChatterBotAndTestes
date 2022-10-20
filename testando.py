from os import chdir, getcwd, listdir
import zipfile
import openpyxl

#cam = input('Digite o caminho: ')
CAMINHO = 'C:/Users/ECOELETRICA/Downloads'
chdir(CAMINHO)
print(getcwd())

for c in listdir():
    try:
        if c.split(".")[1] == 'zip':
            if c.split("_")[0] == "TESTANDO":
                print(c)
                z = zipfile.ZipFile(c,'r')
                z.extractall('C:/Users/ECOELETRICA/Desktop')
                z.close()
    except:
       print('Erro na extração')

CAMINHO = 'C:/Users/ECOELETRICA/Downloads'
chdir(CAMINHO)

for arquivo in listdir():
    try:
        if arquivo.split(".")[1] == 'xlsx':
            DATA = openpyxl.load_workbook(filename=CAMINHO+"/"+arquivo)
            for sheet in DATA.sheetnames:
                for d in DATA[sheet].iter_rows(values_only=True):
                    print(d)
    except:
        print("erro na leitura do excell")